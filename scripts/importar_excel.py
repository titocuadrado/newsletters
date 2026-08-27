#!/usr/bin/env python3
"""
Importa los Excel de planning de newsletters a los ficheros de datos del planificador.

Acepta los dos formatos que ha usado Marketing:

  FORMATO RICO (Planning_Newsletters_V18, 2022-2024) — el mejor de los dos:
      DIA | TEMA | PRODUCTOS | BLOG | ENLACES (hasta 5 columnas) | DISCLAIMERS
      El TEMA es un gancho editorial ("Cafeteros", "Tiempo de sopa", "Burger lovers")
      separado de los PRODUCTOS que entran en el envio. Los DISCLAIMERS son avisos
      de feria por mercado ("EXPORT: SIGEP / FRANCIA: EUROPAIN").

  FORMATO SIMPLE (Planning_NL_*, 2024-2026):
      fecha | tema | alternativa | estado
      Todo dentro de la celda del tema: asunto, paginas de catalogo y notas internas.

El anyo no esta escrito en ninguna parte: se infiere probando anyos de arranque y
quedandose con el que hace cuadrar los dias de la semana declarados.

Los ficheros se pasan de MAS ANTIGUO a MAS NUEVO. Cuando dos ficheros planifican
la misma fecha con temas distintos (una hoja revisada sobre otra), gana el fichero
mas nuevo y el anterior queda marcado como 'revision_previa': no cuenta como envio.

Uso:  python3 scripts/importar_excel.py data/origen-Planning_Newsletters_V18.xlsx \
                                        data/origen-Planning_NL_primaveraestiu.xlsx \
                                        data/origen-Planning_NL_2025.xlsx \
                                        data/origen-Planning_NL_2026.xlsx
"""
import openpyxl, re, json, sys, unicodedata, argparse, os
from datetime import date
from collections import Counter, defaultdict
from clasificar import fabricacion

MESES = {'enero':1,'febrero':2,'marzo':3,'abril':4,'mayo':5,'junio':6,'julio':7,'agosto':8,
         'septiembre':9,'setiembre':9,'octubre':10,'noviembre':11,'diciembre':12,
         'ene':1,'feb':2,'mar':3,'abr':4,'may':5,'jun':6,'jul':7,'ago':8,'sep':9,'oct':10,
         'nov':11,'dic':12}
DIAS = {'lunes':0,'martes':1,'miercoles':2,'jueves':3,'viernes':4,'sabado':5,'domingo':6,
        'juves':3,'lun':0,'mar':1,'mie':2,'jue':3,'vie':4}
NOMBRE_DIA = ['lunes','martes','miercoles','jueves','viernes','sabado','domingo']
TEMPORADAS = {12:'invierno',1:'invierno',2:'invierno',3:'primavera',4:'primavera',5:'primavera',
              6:'verano',7:'verano',8:'verano',9:'otono',10:'otono',11:'otono'}

def sinacentos(s):
    s = unicodedata.normalize('NFD', str(s).lower())
    return ''.join(c for c in s if unicodedata.category(c) != 'Mn')

def slug(s):
    return re.sub(r'[^a-z0-9]+', '-', sinacentos(s)).strip('-')[:60]

def parse_fecha(txt):
    """'Martes 18 junio' / 'Mar 6 oct' / 'jueves 10 de julio' -> (dow|None, dia, mes)"""
    t = re.sub(r'\s+', ' ', sinacentos(txt)).strip().replace(' de ', ' ')
    m = re.match(r'^([a-z]+)\.?\s+(\d{1,2})\s+([a-z]+)', t)
    if not m:
        m = re.match(r'^(\d{1,2})\s+([a-z]+)', t)          # "6 oct" sin dia de semana
        if not m or m.group(2) not in MESES:
            return None
        return None, int(m.group(1)), MESES[m.group(2)]
    mes = MESES.get(m.group(3))
    return (DIAS.get(m.group(1)), int(m.group(2)), mes) if mes else None

# ── Deteccion de formato ────────────────────────────────────────────────────
def es_formato_rico(ws):
    cab = [sinacentos(c.value or '') for c in ws[1]]
    return 'dia' in cab and 'tema' in cab

def columnas_ricas(ws):
    """Localiza las columnas por su cabecera. ENLACES ocupa varias sin cabecera propia."""
    cab = [sinacentos(c.value or '') for c in ws[1]]
    idx = {'fecha': 0, 'tema': 1}
    for nombre, clave in (('productos','productos'), ('blog','blog'),
                          ('enlaces','enlaces'), ('disclaimers','disclaimers')):
        if clave in cab:
            idx[nombre] = cab.index(clave)
    fin_enlaces = idx.get('disclaimers', len(cab))
    idx['enlaces_rango'] = (idx.get('enlaces', 4), fin_enlaces)
    return idx

# ── Extraccion de campos del texto libre ────────────────────────────────────
def paginas_de(texto):
    pags = set()
    for m in re.finditer(r'p[ag]{1,3}(?:ina)?s?\.?\s*([\d\s,\-y]+)', texto, re.I):
        pags.update(int(n) for n in re.findall(r'\d{1,3}', m.group(1)))
    for m in re.finditer(r'\((\d{1,3}(?:\s*[-,]\s*\d{1,3})*)\)', texto):
        pags.update(int(n) for n in re.findall(r'\d{1,3}', m.group(1)))
    return sorted(p for p in pags if 1 <= p <= 600)

MARCAS_PROCESO = ['(feta)','feta)','repetida','repetir','rep )','rep)','(rep','no stock',
                  'resend','dos parts','dos enllacos','nomes els de dalt','hi ha','esmenta',
                  'per exemple','una mica de tot','les diferents linies','per tornar',
                  'aquests de cara','nous colors','nous dissenys','no hay materia prima',
                  'amb nova tarifa','ojo,','clientes web','cllientes web']

def notas_de(texto):
    t = sinacentos(texto)
    return sorted({m for m in MARCAS_PROCESO if m in t})

def numero_de(tema):
    """'690 - Airlaid Natural' -> 690. Numeracion de campanya del formato antiguo."""
    m = re.match(r'^(\d{3})\s*[-–]\s*', tema.strip())
    return int(m.group(1)) if m else None

def limpia_tema(tema):
    t = re.sub(r'https?://\S+', '', tema)
    t = re.sub(r'^\d{3}\s*[-–]\s*', '', t.strip())         # quita el numero de campanya
    return re.sub(r'\s{2,}', ' ', t).strip(' .·-+\n')

TIPOS = [
    ('lanzamiento-catalogo', ['catalogo','cataleg','flip_outlet']),
    ('tarifa',               ['tarifa','tarifas','bajada de precio','actualizacion tarifas']),
    ('promocion',            ['black friday','black november','outlet','promo']),
    ('institucional',        ['felicitacion','encuesta','web print','historia portadas',
                              'nuestras colecciones','presentacio','suite','digital print']),
    ('contenido-blog',       ['blog']),
    ('video',                ['video']),
]
def tipo_de(tema, blog, notas):
    t = sinacentos(tema)
    for nombre, claves in TIPOS:
        if any(k in t for k in claves):
            return nombre
    if blog:
        return 'producto-con-blog'
    if any(k in notas for k in ('repetida','repetir','(rep','rep)','rep )','resend')):
        return 'repeticion'
    if re.search(r'\bnew\b|\bnuev[oa]s?\b|novedad', t):
        return 'novedad'
    return 'producto'

FAMILIAS = [
    ('institucional', ['catalogo','cataleg','tarifa','felicitacion','encuesta','web print',
                       'historia portadas','nuestras colecciones','normativa','presentacio',
                       'digital print','black friday','black november','outlet']),
    ('cotillon',      ['navidad','cotillon','kerstmis','christmas','fin de ano','san valentin',
                       'lotes','regalo','conjuntos decorados','finger food']),
    ('take-away',     ['take away','take-away','delivery','grab','hamburguesa','burger','hot dog',
                       'kebab','wrap','crepe','sushi','bagel','poke','taco','street food','pollo',
                       'ensaladera','barquilla','barquita','maletas','fritos','palomitas',
                       'cucurucho','papelina','pizza','panini','sandwich','bocadillo','lunch box',
                       'noodles','comida oriental','cajetilla','concha','paper lock','portavasos',
                       'transporte para vasos','cajas de comida','cajas para comida','salsas']),
    ('un-solo-uso',   ['servilleta','mantel','mantelin','tete','camino de mesa','bolsa','vaso',
                       'pajita','canutillo','paletina','agitador','envase','recipiente','tarrina',
                       'blonda','posavaso','bandeja','papel antigrasa','antigrasas','pasteleria',
                       'aluminio','film','cubierto','plato','bol','cucharilla','pincho','picks',
                       'toallita','baberos','contenedores lacados','sobres adhesivos','aros de horneado',
                       'molde','capsula','petits fours','petit fours','muffin','panettone','deco',
                       'like linen','like-linen','airlaid','spunbond','spundbond','double point',
                       'just in time','mini servis','maxi servis','canguro','cangurito','kangaroo',
                       'dry cotton','thepack','the pack','bionic','wood','paper cutlery','paper spoon',
                       'parole','times','pfas','feel green','areca','just paper','fitipaldi','leaf',
                       'cera vegetal','compostable','plastic free','pla y cpla','rpet','entretenimiento',
                       'colorear','etiqueta','looks','keiko','safari','jazz','ipanema','vichy',
                       'tartan','dakar','scottish','reciclad','fsc','eco bamboo','eco bambu']),
    ('bar-buffet',    ['bufet','buffet','presentador','display','chafing','dispensador','cubitera',
                       'champanera','jarra','termo','barman','coctel','exprimidor','ramequin',
                       'mini recipiente','aperitivo','bambu','tabla','botellero','estanteria',
                       'american style','artinox','waki','cristaleria','cristal','loza','lechera',
                       'expositor','baskets','storage','mini platos','mini vasos','bar y coctel']),
    ('mesa-accesorios',['cuberteria','cubiertos inox','porcelana','melamina','policarbonato',
                       'stoneware','akala','asamiware','enamelware','vajilla','copa','irrompible',
                       'reutilizable','reuse it','cesta','servilletero','lampara','vela','mobiliario',
                       'mesa plegable','silla','taburete','trona','set lounge','oslo','essenza',
                       'dakota','olivia','siena','atlanta','coral','lyon','sevilla','spiga',
                       'marlene','provenza','vintage','supratek','supra-tek','chicago','accesorios de mesa']),
    ('utensilios-cocina',['bateria','cuchillo','cubeta','contenedor','isotermic','bolsa termica',
                       'rack','carro','tabla de corte','bascula','termometro','maquinaria',
                       'termosellar','termosellable','mesas de trabajo','utensilios cocina','pala']),
    ('limpieza-higiene',['celulosa','papel higienico','secamanos','bobina','matatrapos','bayeta',
                       'pano','basura','papelera','reciclaje','guante','vestuario','delantal',
                       'gorro','jabon','microkleen','perfokleen','flushable','roll it','cubos',
                       'proteccion e higiene','limpieza','panuelos']),
    ('habitaciones',  ['amenitie','amenities','gel','champu','acogida','dental','afeitado',
                       'costurero','zapatilla','toalla','percha','reposamaletas','spa','allure',
                       'aloe vera','azur','therapy','touch of charm','toiletries','habitacion']),
    ('rotulacion',    ['portamenu','pizarra','comanda','senalizacion','catenaria','porta cuenta',
                       'pulsera','papel termico','tpv','rotulacion']),
]
def familia_de(texto):
    t = sinacentos(texto)
    for fam, claves in FAMILIAS:
        if any(k in t for k in claves):
            return fam
    return 'sin-clasificar'

MARCAS = ['THEPACK','BIONIC','FEEL GREEN','PFAS FREE','PAROLE','TIMES','WOOD','DRY COTTON',
          'Like-Linen','Like Linen','Airlaid','Spunbond','Double Point','Just in Time','Mini Servis',
          'Kangaroo','Cangurito','ERIK','Ipanema','Jazz','Keiko','Safari','Vichy','Akala','Artinox',
          'Enamelware','Asamiware','Crockery','American Style','Waki Glass','Wasara','Fitipaldi',
          'Leaf','Areca','Vintage','Kintsu','Supratek','Arrow','Paper Cutlery','Paper Spoon',
          'Paper Lock','Just Paper','Reuse it','Grill & Go','Roast and Go','Ovensafe','CUBE',
          'Set Lounge','Flurry','Grafitti','Microkleen','Perfokleen','Push Pack','Atlanta','Coral',
          'Dakota','Essenza','Lyon','Marlene','Olivia','Oslo','Provenza','Sevilla','Siena','Spiga',
          'Allure','Aloe Vera','Azur','Therapy','Touch of Charm','Chicago Lava Stone','FSC']
def marcas_de(texto):
    t = sinacentos(texto)
    return sorted({m for m in MARCAS if sinacentos(m) in t})

# ── Lectura de una hoja ─────────────────────────────────────────────────────
def filas_crudas(ws):
    """Normaliza cualquiera de los dos formatos a un dict comun."""
    rico = es_formato_rico(ws)
    idx = columnas_ricas(ws) if rico else None
    salida = []
    for i, r in enumerate(ws.iter_rows(values_only=True), start=1):
        cel = ['' if c is None else str(c).strip() for c in r] + [''] * 12
        if rico and i == 1:
            continue
        if not any(cel):
            continue
        if rico:
            fecha_raw = cel[idx['fecha']]
            tema      = cel[idx['tema']]
            productos = cel[idx.get('productos', 2)]
            blog      = cel[idx.get('blog', 3)]
            a, b      = idx['enlaces_rango']
            enlaces   = [c for c in cel[a:b] if c.startswith('http')]
            disclaimer = cel[idx.get('disclaimers', 8)]
            alternativa = ''
        else:
            fecha_raw, tema = cel[0], cel[1]
            productos = blog = disclaimer = ''
            resto = ' '.join(x for x in (cel[2], cel[3]) if x)
            enlaces = re.findall(r'https?://\S+', tema + ' ' + resto)
            alternativa = re.sub(r'https?://\S+', '', resto).strip()
        if not tema:
            continue
        if 'PLANNING' in tema.upper() and not fecha_raw:
            continue
        salida.append({'fila': i, 'fecha_raw': fecha_raw, 'tema_raw': tema,
                       'productos': productos, 'blog': blog, 'enlaces': enlaces,
                       'disclaimer': disclaimer, 'alternativa': alternativa})
    return salida

def fechar(filas, anyo_inicial):
    """Asigna anyo recorriendo en orden; incrementa al saltar de diciembre a enero.
    Devuelve (lista_fechadas, n_incoherencias)."""
    anyo, mes_previo, out, malos = anyo_inicial, None, [], 0
    for f in filas:
        p = parse_fecha(f['fecha_raw']) if f['fecha_raw'] else None
        if not p:
            out.append((f, None)); continue
        dow, dia, mes = p
        if mes_previo is not None and mes < mes_previo:
            anyo += 1
        mes_previo = mes
        try:
            d = date(anyo, mes, dia)
        except ValueError:
            out.append((f, None)); malos += 1; continue
        if dow is not None and d.weekday() != dow:
            malos += 1
        out.append((f, d))
    return out, malos

def infiere_anyo(filas, candidatos=range(2020, 2028)):
    """Prueba anyos de arranque y se queda con el de menos incoherencias de dia."""
    mejor, mejor_malos = None, None
    for a in candidatos:
        _, malos = fechar(filas, a)
        if mejor_malos is None or malos < mejor_malos:
            mejor, mejor_malos = a, malos
        if malos == 0:
            break
    return mejor, mejor_malos

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('xlsx', nargs='+')
    args = ap.parse_args()

    envios, ideas, avisos = [], [], []

    for autoridad, ruta in enumerate(args.xlsx):
        wb = openpyxl.load_workbook(ruta, data_only=True)
        origen_fichero = os.path.basename(ruta)
        for ws in wb.worksheets:
            filas = filas_crudas(ws)
            if not filas:
                continue
            anyo, malos = infiere_anyo(filas)
            if malos:
                avisos.append(f'{origen_fichero}!{ws.title}: {malos} dias de semana no cuadran '
                              f'con el anyo inferido ({anyo})')
            fechadas, _ = fechar(filas, anyo)
            for f, d in fechadas:
                texto = ' '.join([f['tema_raw'], f['productos'], f['alternativa']])
                tema = limpia_tema(f['tema_raw'])
                reg = {
                    'asunto': tema,
                    'productos': f['productos'],
                    'blog': f['blog'],
                    'enlaces': f['enlaces'],
                    'disclaimer': f['disclaimer'],
                    'alternativa': f['alternativa'],
                    'numero': numero_de(f['tema_raw']),
                    'familia': familia_de(texto),
                    'fabricacion': fabricacion(texto),
                    'marcas': marcas_de(texto),
                    'paginas_catalogo': paginas_de(texto),
                    'notas_internas': notas_de(texto + ' ' + f['tema_raw']),
                    'formato_origen': 'rico' if es_formato_rico(ws) else 'simple',
                    'autoridad': autoridad,
                    'origen': f'{origen_fichero}!{ws.title}!{f["fila"]}',
                }
                reg['tipo'] = tipo_de(f['tema_raw'], f['blog'], reg['notas_internas'])
                if d:
                    reg.update(fecha=d.isoformat(), hora='15:00',
                               dia_semana=NOMBRE_DIA[d.weekday()],
                               temporada=TEMPORADAS[d.month], mes=d.month, anyo=d.year)
                    envios.append(reg)
                else:
                    ideas.append(reg)

    # ── Fusion: misma fecha + mismo asunto = el mismo envio visto en dos ficheros.
    #    Gana el registro con mas informacion (el formato rico la tiene toda).
    def riqueza(r):
        return (len(r['enlaces']), bool(r['productos']), bool(r['blog']),
                bool(r['disclaimer']), len(r['paginas_catalogo']))
    fusion = {}
    for r in envios:
        k = (r['fecha'], slug(r['asunto']))
        if k not in fusion or riqueza(r) > riqueza(fusion[k]):
            fusion[k] = r
    envios = sorted(fusion.values(), key=lambda e: e['fecha'])

    # Misma fecha con asuntos distintos = una hoja revisada sobre otra.
    # Gana el fichero mas nuevo; los demas quedan como revision previa y no cuentan.
    por_fecha = defaultdict(list)
    for r in envios:
        por_fecha[r['fecha']].append(r)
    n_revisiones = 0
    for rs in por_fecha.values():
        if len(rs) == 1:
            continue
        rs.sort(key=lambda r: (r['autoridad'], riqueza(r)), reverse=True)
        for r in rs[1:]:
            r['revision_previa'] = True
            n_revisiones += 1
    reales = [r for r in envios if not r.get('revision_previa')]

    # ── Banco de temas ──────────────────────────────────────────────────────
    banco = {}
    for reg in reales + [r for r in envios if r.get('revision_previa')] + ideas:
        if not reg['asunto']:
            continue
        base = re.sub(r'^(repetida|repetir|resend)\s*:?\s*', '', sinacentos(reg['asunto'])).strip()
        base = re.sub(r'\(rep[^)]*\)|\(repetida\)', '', base).strip()
        k = slug(base)
        if not k:
            continue
        t = banco.setdefault(k, {
            'id': k, 'nombre': reg['asunto'], 'familia': reg['familia'],
            'fabricacion': reg['fabricacion'], 'tipo': reg['tipo'],
            'marcas': [], 'paginas_catalogo': [], 'enlaces': [], 'productos': '',
            'blog': '', 'notas_internas': [], 'envios': [], '_meses': [],
        })
        t['marcas'] = sorted(set(t['marcas']) | set(reg['marcas']))
        t['paginas_catalogo'] = sorted(set(t['paginas_catalogo']) | set(reg['paginas_catalogo']))
        t['enlaces'] = sorted(set(t['enlaces']) | set(reg['enlaces']))
        t['notas_internas'] = sorted(set(t['notas_internas']) | set(reg['notas_internas']))
        if len(reg['productos']) > len(t['productos']):
            t['productos'] = reg['productos']
        if reg['blog'] and not t['blog']:
            t['blog'] = reg['blog']
        if reg.get('fecha') and not reg.get('revision_previa'):
            t['envios'].append(reg['fecha'])
            t['_meses'].append(reg['mes'])
        if t['familia'] == 'sin-clasificar' != reg['familia']:
            t['familia'] = reg['familia']
        if t['fabricacion'] == 'por-confirmar' != reg['fabricacion']:
            t['fabricacion'] = reg['fabricacion']

    for t in banco.values():
        t['envios'].sort()
        t['veces_enviado'] = len(t['envios'])
        t['ultimo_envio'] = t['envios'][-1] if t['envios'] else None
        t['temporadas'] = sorted({TEMPORADAS[m] for m in t['_meses']})
        t['meses_usados'] = sorted(set(t['_meses']))
        del t['_meses']
    temas = sorted(banco.values(), key=lambda t: (-t['veces_enviado'], t['id']))

    os.makedirs('data', exist_ok=True)
    json.dump({'origen': [os.path.basename(x) for x in args.xlsx], 'envios': envios},
              open('data/historico.json', 'w'), ensure_ascii=False, indent=1)
    json.dump({'temas': temas}, open('data/temas.json', 'w'), ensure_ascii=False, indent=1)

    print(f'ficheros ................. {len(args.xlsx)}')
    print(f'envios fechados .......... {len(reales)}  ({reales[0]["fecha"]} -> {reales[-1]["fecha"]})')
    print(f'revisiones descartadas ... {n_revisiones}')
    print(f'ideas sin fecha .......... {len(ideas)}')
    print(f'temas unicos ............. {len(temas)}')
    print(f'  con enlaces ............ {sum(1 for t in temas if t["enlaces"])}')
    print(f'  con productos detallados {sum(1 for t in temas if t["productos"])}')
    print(f'  con blog asociado ...... {sum(1 for t in temas if t["blog"])}')

    print(f'avisos ................... {len(avisos)}')
    for a in avisos:
        print('   !', a)
    print('\nenvios por anyo:')
    for a, c in sorted(Counter(e['anyo'] for e in reales).items()):
        print(f'   {a}  {c:4d}')
    print('\norigen de fabricacion (envios):')
    for f, c in Counter(e['fabricacion'] for e in reales).most_common():
        print(f'   {c:4d}  {c/len(reales)*100:5.1f}%  {f}')
    print('\nfamilia (envios):')
    for f, c in Counter(e['familia'] for e in reales).most_common():
        print(f'   {c:4d}  {c/len(reales)*100:5.1f}%  {f}')

if __name__ == '__main__':
    main()
